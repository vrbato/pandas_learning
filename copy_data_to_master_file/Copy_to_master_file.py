#read values from closed excel files
from pathlib import Path
from openpyxl import load_workbook, Workbook

#1. KROK
SOURCE_DIR = "Daily_Reports"
excel_files = list(Path(SOURCE_DIR).glob("*xlsx"))

#2.KROK - procházení jednotlivých souborů a vytažení jejich dat do listu a uložení do slovníku
values_excel_files = {}
for excel_file in excel_files:
    report_date = excel_file.stem.replace("_Report","") #odstraní to z názvu souboru _Rerport.xlsx a bude to sloužit jako klíče do slovníku
    #otevření každého souboru, který prochází a vytažení hodnot
    wb = load_workbook(filename=excel_file) #otvírá to soubor
    rng = wb["Sheet1"]["B2":"B19"]
    rng_values = []
    for cells in rng: #tohle vyhodí celej tuple hodnot - musím vnořit další FOR pro jednotlivou buňku
        for cell in cells:
            rng_values.append(cell.value) #naháže mi to do listu hodnoty
        values_excel_files[report_date] = rng_values #klíč jsou datumy ze souboru a k tomu přidávám hodnoty z každého souboru, které jsou uloženy do listu

# print(len(values_excel_files))
# print(values_excel_files)

#3.KROK - ukládání do MASTER SOUBORU
wb = load_workbook(filename="Masterfile_Template.xlsx")
for ws in wb.worksheets:
    clm = "B"
    first_row = 3
    last_row = len(ws[clm]) #počet řádků může být rozdílný v souborech
    rng = ws[f"{clm}{first_row}:{clm}{last_row}"] #určím si pomocnou range na listu
    for cells in rng:
        for cell in cells:
            if cell.value in values_excel_files:
                for i, value in enumerate(values_excel_files[cell.value]):
                    cell.offset(row=0, column=i + 1).value = value
wb.save("Masterfile_filledup_muj.xlsx")
